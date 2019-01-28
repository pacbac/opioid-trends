from openpyxl import load_workbook
import os

""" Compare similarity in results between markov and county rank models """

ROOT_WDIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
MARKOV_DIR = f'{ROOT_WDIR}/generated_data/DrugMarkov.xlsx'
CRANK_DIR = f'{ROOT_WDIR}/generated_data/countyRank_model_state_origin.xlsx'
markov = load_workbook(MARKOV_DIR).active
crank = load_workbook(CRANK_DIR).active

notEqual = 0
total = 0
for j in range(1, 6):
    c = 3*j
    for i in range(1, 69):
        crankval = crank.cell(row=i, column=c).value.upper().split(" ")[0]
        print(crankval)
        markovval = markov.cell(row=i, column=c).value.upper().split(" ")[0]
        if crankval != markovval:
            notEqual += 1
        total += 1

print(f'{notEqual}/{total} total results are different')