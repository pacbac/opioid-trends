import os
from openpyxl import load_workbook, Workbook

"""
User input: State, County, Drug
Output: Drug identification count for 2010 to 2017 (if available)
"""
ROOT_WDIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
DATA_DIR = f'{ROOT_WDIR}/2019_MCMProblemC_DATA/MCM_NFLIS_Data.xlsx'
wb = load_workbook(DATA_DIR)
sheet = wb.worksheets[1]
states = {'KY': {}, 'VA': {}, 'WV': {}, 'PA': {}, 'OH': {}} # load drug counts in each state
counties = {'KY': [], 'VA': [], 'WV': [], 'PA': [], 'OH': []} # counties list per state

yrs = [ yr for yr in range(2010, 2018) ]

def loadDrugCount():
    global states

    for j in states.keys():
        states[j] = {}
        for yr in yrs:
            states[j][yr] = {}

    for i in range(2, sheet.max_row+1):
        state = sheet.cell(row=i, column=2).value
        county = sheet.cell(row=i, column=3).value
        yr = sheet.cell(row=i, column=1).value
        drug = sheet.cell(row=i, column=7).value
        drugCount = sheet.cell(row=i, column=8).value
        tbl = states[state]
        if (county, drug) not in tbl[yr]: # new county, add it to records
            tbl[yr][(county, drug)] = drugCount

def search(stateIn, countyIn, drugIn):
    global states
    result = {}
    for yr in yrs:
        if (countyIn, drugIn) in states[stateIn][yr]:
            result[yr] = states[stateIn][yr][(countyIn, drugIn)]
    return result

def loadCountyList():
    for i in range(2, sheet.max_row+1):
        county = sheet.cell(row=i, column=3).value
        state = sheet.cell(row=i, column=2).value
        if county not in counties[state]:
            counties[state].append(county)

def load():
    loadDrugCount()
    loadCountyList()

def main():
    global states
    print("Load from NFLIS complete.")
    while True:
        print("Type in State,County,SubstanceName (separate by comma):")
        userIn = input()
        stateIn, countyIn, drugIn = tuple(userIn.split(","))
        print(f'Results for {drugIn} in {countyIn}, {stateIn}:')
        for yr, drugCount in search(stateIn, countyIn, drugIn).items():
            print(f'  {yr}:', drugCount)

load()
if __name__ == "__main__":
    main()