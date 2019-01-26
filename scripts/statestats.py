import os
from openpyxl import load_workbook, Workbook

DATA_DIR = f'{os.getcwd()}/2019_MCMProblemC_DATA/MCM_NFLIS_Data.xlsx'
wb = load_workbook(DATA_DIR)
sheet = wb.worksheets[1]

yrToCol = {
    2010: 'a',
    2011: 'b',
    2012: 'c',
    2013: 'd',
    2014: 'e',
    2015: 'f',
    2016: 'g',
    2017: 'h'
}

def main():
    states = {'KY': {}, 'VA': {}, 'WV': {}, 'PA': {}, 'OH': {}}
    curYr = 2010
    statesWb = {'KY': Workbook(), 'VA': Workbook(), 'WV': Workbook(), 'PA': Workbook(), 'OH': Workbook()}
    print(sheet.cell(row=1, column=1).value, sheet.cell(row=1, column=2).value, sheet.cell(row=1, column=3).value)

    for j in statesWb.keys():
        newSheet = statesWb[j].active
        newSheet.cell(row=1, column=1).value = 'County'
        for yr in yrToCol.keys():
            newSheet.cell(row=1, column=(yr-2008)).value = yr
        states[j]['counties'] = []
        states[j]['entries'] = {}
        for yr in yrToCol.keys():
            states[j]['entries'][yr] = []

    prevCounty = None
    for i in range(2, sheet.max_row+1):
        state = sheet.cell(row=i, column=2).value
        county = sheet.cell(row=i, column=3).value
        yr = sheet.cell(row=i, column=1).value
        drugCount = sheet.cell(row=i, column=9).value
        newSheet = statesWb[state].active
        tbl = states[state]
        if county not in tbl['counties']: # new county, add it to records
            tbl['counties'].append(county)
            tbl['entries'][yr].append((county, drugCount))
        elif prevCounty != county: # skip over repeating county entries
            tbl['entries'][yr].append((county, drugCount))
        prevCounty = county
    
    print(len(states['VA']['counties'])) 

    for state in states.keys():
        newSheet = statesWb[state].active
        counties = states[state]['counties']
        entries = states[state]['entries']
        for i, county in enumerate(counties, 2):
            newSheet.cell(row=i, column=1).value = county

        for yr in entries.keys():
            for county, drugCount in entries[yr]:
                newSheet.cell(row=counties.index(county)+2, column=(yr-2008)).value = drugCount
        statesWb[state].save(f'generated_data/{state}test.xlsx')


if __name__ == "__main__":
    main()