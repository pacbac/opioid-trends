import os, sys
# from openpyxl import load_workbook
ROOT_WDIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.insert(0, ROOT_WDIR)
from scripts.searchDrugCounty import search
from Classes.county import County, getAllCounties, allCounties as counties
import networkx as nx
import numpy as np
from random import random
from openpyxl import load_workbook

DATA_DIR = f'{ROOT_WDIR}/generated_data/DrugMarkov.xlsx'
wb = load_workbook(DATA_DIR)
sheet = wb.active

"""
1. Find earliest year that counties use a drug
2. The resulting counties in 2010 are nodes Vn = count of drug identifications that year
3. p_ij = cnt_i/( (total cnt across all nodes)(1 + distance_ij) )
4. p_ij' = unify p_ij so that 
5. Each round calculate which people stay (from A to A) and which people move (from A to B, from A to C, ...., etc)
6. If pop A = 0 (A dies), remove it 
"""
BEGIN_YR = 2017
yrs = [yr for yr in range(BEGIN_YR, 2018)]
stateToCol = {
    'VA': 3,
    'OH': 6,
    'KY': 9,
    'PA': 12,
    'WV': 15
}
ELIMINATE_THRESHOLD = 0
SIM_ITERS = 30

getAllCounties()
""" 
Given a drug, return the counties that used this drug in BEGIN_YR
These counties are candidates as the source of the drug
"""
def getCountyList(drug):
    # get counties that began using the drug in 2010 (starting nodes in the graph)
    global BEGIN_YR
    BEGIN_YR = 2017
    countyList = []
    for county in counties:
        searchResult = search(county.m_state, county.m_name, drug)
        if searchResult:
            startYr = min(searchResult.keys())
            if BEGIN_YR > startYr:
                BEGIN_YR = startYr
            countyList.append((county, startYr)) # get smallest key in dict result (starting year)
    # keep only candidate counties who started using this drug in BEGIN_YR (beginning)
    return list(map(lambda tup: [tup[0], tup[0].drugCases(drug, BEGIN_YR)], list(filter(lambda tup: tup[1] == BEGIN_YR, countyList))))

def calcEdge(dCase, totalCases, dist):
    return float(dCase)/(totalCases * dist)

def normalizeEdges(G):
    sumWeight = 0
    for node_i, node_j, d in G.edges(data=True):
            sumWeight += d['weight'] # float

    for node_i, node_j, d in G.edges(data=True):
        d['weight'] /= float(sumWeight)    

def initSim(countyList, drug):
    G = nx.DiGraph()
    # add candidate counties to graph
    for county in countyList:
        G.add_node(county)
    # add relationships between counties
    totalCases = County.sumPop(lambda x: x.drugCases(drug, BEGIN_YR), drug, BEGIN_YR, countyList)
    for county_i in countyList:
        for county_j in countyList:
            drugcases_j = county_j.drugCases(drug, BEGIN_YR)
            dist_ij = 1 + county_i.distanceTo(county_j)
            G.add_edge(county_i, county_j, weight=calcEdge(drugcases_j, totalCases, dist_ij))
    normalizeEdges(G)
    return G

def main(drug, drugNum):
    origCountyListCases = getCountyList(drug)
    origCountyList = list(map(lambda tup: tup[0], origCountyListCases))
    itList = list(map(lambda tup: tup[0].m_name, origCountyListCases))
    itCount = [0 for it in itList]
    stateRecords = {}
    for county in origCountyList:
        if county.m_state not in stateRecords:
            stateRecords[county.m_state] = {}
        stateRecords[county.m_state][county.m_name] = 0

    for it in range(0, SIM_ITERS):
        print("Round", it, "start.")
        countyListCases = getCountyList(drug)
        maxCase = max(list(map(lambda tup: tup[1], countyListCases))) # highest drug case amount of all counties
        ELIMINATE_THRESHOLD = 0.4*maxCase if maxCase > 25 else 0
        # filter out counties with drug cases that are sufficiently small (negligible)
        countyListCases = list(filter(lambda tup: tup[1] > ELIMINATE_THRESHOLD, countyListCases))
        countyList = list(map(lambda tup: tup[0], countyListCases))
        stateCount = {}
        for county in countyList:
            stateCount[county.m_state] = 0 if county.m_state not in stateCount else stateCount[county.m_state] + 1
        G = initSim(countyList, drug)

        while(len(G) > 1):
            # go through each county
            for i, countyCase in enumerate(countyListCases):
                [county_i, dCases] = countyCase
                for k in range(0, dCases): # go through each "case"
                    for j, county_j in enumerate(countyList): 
                        if random() < G[county_i][county_j]['weight']: # if "case" decides to move
                            countyListCases[i][1] -= 1 # decrease drug case amount in host county
                            countyListCases[j][1] += 1 # increase drug case amount in neighboring county

            # remove any dead nodes from 
            i = 0
            while i < len(countyListCases):
                county_i, dCases = countyListCases[i]
                if dCases <= 0: 
                    stateCount[county_i.m_state] -= 1
                    if stateCount[county_i.m_state] <= 0:
                        stateRecords[county_i.m_state][county_i.m_name] += 1
                    G.remove_node(county_i)
                    countyListCases.pop(i)
                    countyList.pop(i)
                else:
                    i += 1

            # refresh node weights
            totalCases = County.sumPop(lambda x: x.drugCases(drug, BEGIN_YR), drug, BEGIN_YR, countyList)
            for county_i in countyList:
                for county_j, dCaseJ in countyListCases:
                    dist_ij = 1 + county_i.distanceTo(county_j)
                    G[county_i][county_j]['weight'] = calcEdge(dCaseJ, totalCases, dist_ij)
            normalizeEdges(G)

        for node in G.nodes:
            stateRecords[node.m_state][node.m_name] += 1
            itCount[itList.index(node.m_name)] += 1

    print("Overall drug source:", itCount[itCount.index(max(itCount))], itList[itCount.index(max(itCount))])
    print("Drug source per state:")
    for state in stateRecords:
        name, count = sorted(list(stateRecords[state].items()), key=lambda kv: kv[1])[-1]
        if count > 0:
            print(f'{state}:', name)
            sheet.cell(row=drugNum, column=stateToCol[state]).value = name
        else:
            county, count = sorted(list(filter(lambda case: case[0].m_state == state, origCountyListCases)), key=lambda kv: kv[1])[-1]
            print(f'{state}:', county.m_name)
            sheet.cell(row=drugNum, column=stateToCol[state]).value = county.m_name
    wb.save(DATA_DIR)

if __name__ == "__main__":
    for i in range(1, sheet.max_row+1):
        drug = sheet.cell(row=i, column=2).value
        print("Start sim", drug)
        main(drug, i)
        print("simulation completed.")